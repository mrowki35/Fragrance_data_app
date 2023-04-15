from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import requests
from datetime import date 

def notino_datagetter(x,main_window):
    if(x!=''):
        url = x
        try:
            result=requests.get(url)
        except requests.exceptions.RequestException:
            main_window.notino_url.background_color = (1, 0, 0, 0.1)
            return 0
        doc = BeautifulSoup(result.text, "html.parser")

        price_tag = doc.find("span",{"class":"sc-1ctnxgp-0 jIlnHX"})
        if price_tag != None:
            value_tag = price_tag.find("span") 
            price = value_tag.text.replace(",",".")
        else:
            prices = doc.find_all(id="pd-price")
            parent=prices[0].parent
            span = parent.find_all("span")
            price = span[1].text.replace(",",".")

        prices = doc.find_all(id="pd-price")
        parent=prices[0].parent
        span = parent.find_all("span")
        volume = span[0].text

        brands = doc.find_all(id="pdHeader")
        brands_parent = brands[0].parent
        stamp = brands_parent.find_all("span")
        genre = stamp[1].text
        brand = stamp[0].text.replace(genre,"")
        category = stamp[2].text
        date_ = date.today().strftime("%d.%m.%Y")

        wb = load_workbook("Parfume_prices_database.xlsx")
        ws = wb.active

        ws.append([brand,genre,category,volume,float(price.replace('\xa0','').strip()),url,date_])
        wb.save("Parfume_prices_database.xlsx")
        main_window.notino_url.background_color=(.5,1,.2,0.1)
    else:
        pass